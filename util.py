import os, sys
import datetime
import color
from os.path import expanduser, exists
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

######################
# Program Name : util
# Date Revised : November 28, 2022
# Description  : This program is a collection of utility
#                tools for the interaction of a to-do list.
# 
# Functions:
#   get_dir_path
#   get_workbook
#   display row
#   check_date
#   get_categories
#   display_categories
#   cls
#
#######################

#######################
# Function name: get_dir_path
# Date Revised : November 14, 2022
# Description  : Identifies Documents directory for
#                Windows or Home directory for Linux
#                and creates an empty to-do list inside.
#######################
def get_dir_path():    

    #Local Variables
    os_type = sys.platform
    home_dir = expanduser("~")

    #IF Operating system is not Windows
    if (os_type != 'win32'):

        #Get path to to-do directory
        todo_path = os.path.join(home_dir, 'Todo-List/')

        #IF to-do directory doesn't exist
        if os.path.isdir(todo_path) == False:

            #Make to-do directory
            os.mkdir(todo_path)
            color.pr_red("NOTICE: To-Do List directory created in 'Home' directory.")

        #Return to-do directory
        return todo_path

    #ELSE
    else:

        #Join home directory path to Documents directory
        docPath = os.path.join(home_dir, "Documents\\")

        #IF to-do directory doesn't exist
        if os.path.isdir(docPath + "ToDo List\\") == False:

            #Make to-do directory
            os.mkdir(docPath + "ToDo List\\")
            color.pr_red("\nNOTICE: To-Do List directory created in 'Documents' directory.\n")

        #Make variable with ToDo List directory
        fullPath = os.path.join(docPath, "ToDo List\\")

        #Return ToDo List directory
        return fullPath

#######################
# Function name: get_workbook
# Date Revised : November 14, 2022
# Description  : Creates workbook for
#                manipulation of a to-do file
#                or activates workbook if one exists.
#######################
def get_workbook():

    #Local Variables
    todo_dir = get_dir_path()
    date_col = "A"
    priority_col = "B"
    status_col = "C"
    category_col = "D"
    description_col = "E"

    #IF a to-do file already exists
    if(exists(todo_dir + 'Notes.xlsx')):
        
        #Load tp-do file
        workbook = load_workbook(todo_dir + 'Notes.xlsx')

    #ELSE
    else:

        #Create a new workbook
        workbook = Workbook()
        worksheet =  workbook.active

        #Name workbook
        worksheet.title = "ToDo List"

        #Set column length
        worksheet.column_dimensions[date_col].width = 12
        worksheet.column_dimensions[priority_col].width = 10
        worksheet.column_dimensions[status_col].width = 10
        worksheet.column_dimensions[category_col].width = 15        
        worksheet.column_dimensions[description_col].width = 50

        #Set headings style
        worksheet[description_col + "1"].font = Font(color = "00800080", bold = True)
        worksheet[priority_col + "1"].font = Font(color = "00800080", bold = True) 
        worksheet[date_col + "1"].font = Font(color = "00800080", bold = True)
        worksheet[category_col + "1"].font = Font(color = "00800080", bold = True) 
        worksheet[status_col + "1"].font = Font(color = "00800080", bold = True) 

        #Set column headings
        worksheet[description_col + "1"] = "Description"
        worksheet[priority_col + "1"] = "Priority"
        worksheet[date_col + "1"] = "Due Date"
        worksheet[status_col + "1"] = "Status"
        worksheet[category_col + "1"] = "Category"

        #Save to-do workbook
        workbook.save(todo_dir + 'Notes.xlsx')

    #Return workbook
    return workbook

#######################
# Function name: display_row
# Date Revised : November 14, 2022
# Description  : Displays the specific row
#                that is being edited.
#######################
def display_row(workbook, row):
    
    #Local Variables
    worksheet = workbook.active     #Worksheet that minipulates to-do file
    spaceString = ""

    #Clear screen
    cls()
    
    #FOR every attribute in the row
    for cell in worksheet[row]:

        #Adjust empty space
        space = 20 - len(str(cell.value))
        for i in range(space):
            spaceString += " "

        #Print attribute
        print (cell.value, end = spaceString)
        spaceString = ""

    #Print line
    print("")


#######################
# Function name: check_date
# Date Revised : November 14, 2022
# Description  : Checks if the entered
#                date is a valid date input.
#######################
def check_date(date_input):
    
    #Local Variables
    output = True       #Output initially set to true

    #TRY
    try:

        #Split input into date attributes
        month, day, year = date_input.split('/')

        #TRY
        try:

            #Turn date into into int
            datetime.datetime(int(year), int(month), int(day))

        #EXCEPT incorrect data type entered
        except ValueError:

            #Set output to false
            output = False

    #EXCEPT incorrect data type entered
    except ValueError:

        #Set output to false
        output = False

    #IF year length is incorrect
    if(output == True and len(year) < 4):

        #Set output to false
        output = False

    #Return output
    return output

#######################
# Function name: get_categories
# Date Revised : November 14, 2022
# Description  : Returns available task categories.
#######################
def get_categories():

    #Add category list to output
    output = ["Analysis",
              "Programming",
              "Algorithms",
              "Testing" ,
              "Debugging",
              "Deploying",
              "Other"]

    #Return output
    return output

#######################
# Function name: display_categories
# Date Revised : November 14, 2022
# Description  : Returns the formatted category list.
#######################
def display_categories():

    #Get list of categories
    categories = get_categories()

    #Display input message
    color.pr_green("Input a Category:")
    index = 1

    #FOR every category
    for category in categories:

        #Adjust index & display category
        if(index < 10):
            color.pr_purple(" " + str(index) + ". " + category)
        else:
            color.pr_purple(str(index) + ". " + category)
        index+=1

#######################
# Function name: cls
# Date Revised : November 14, 2022
# Description  : Clears the current terminal screen.
#######################
def cls():  

    #Clear current screen
    os.system('cls' if os.name == 'nt' else 'clear')
